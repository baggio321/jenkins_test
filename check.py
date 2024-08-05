# coding=utf-8

from odinUtils.office_doc import OfficeDocService
import openpyxl
import os

def get_doc_data(doc_service, doc_id):
    doc_sheets = doc_service.get_latest_doc(doc_id)
    doc_data = {}
    for sheet in doc_sheets:
        sheet_name = sheet.get("title")
        nrows = sheet.get("nrows")
        ncols = sheet.get("ncols")
        sheet_data = {}
        for row in range(nrows):
            for col in range(ncols):
                cell_value = sheet["data"].get("{0},{1}".format(row, col), "")
                sheet_data[(row, col)] = cell_value
        doc_data[sheet_name] = sheet_data
    return doc_data


def compare_docs(local_data, updated_data):
    new_entries = {}
    for sheet_name, updated_sheet_data in updated_data.items():
        local_sheet_data = local_data.get(sheet_name, {})
        for cell_pos, cell_value in updated_sheet_data.items():
            if local_sheet_data.get(cell_pos) != cell_value:
                if sheet_name not in new_entries:
                    new_entries[sheet_name] = {}
                new_entries[sheet_name][cell_pos] = cell_value
    return new_entries

def create_workbook(file_path):
    workbook = openpyxl.Workbook()
    workbook.save(file_path)

def update_local_excel(local_file_path, new_entries):
    if not os.path.exists(local_file_path):
        create_workbook(local_file_path)

    workbook = openpyxl.load_workbook(local_file_path)
    for sheet_name, sheet_data in new_entries.items():
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)

        for cell_pos, cell_value in sheet_data.items():
            row, col = cell_pos
            sheet.cell(row=row+1, column=col+1, value=cell_value)
    workbook.save(local_file_path)

def main(local_doc_id, updated_doc_id, local_file_path):
    doc_service = OfficeDocService('SESSION=MmVmMzZkNzYtODVmOS00NzMyLTljNzMtYTI1NWMxZmNiOTEx')
    
    local_data = get_doc_data(doc_service, local_doc_id)
    updated_data = get_doc_data(doc_service, updated_doc_id)
    
    new_entries = compare_docs(local_data, updated_data)
    
    update_local_excel(local_file_path, new_entries)


if __name__ == '__main__':
    main('ccdde9f2474b48658aa92f57d3321e0d', '6f7abc192b094c179e297bcd7631bb22', './ECA.xlsx')